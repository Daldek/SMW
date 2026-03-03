"""Tests for ExcelProvider file structure validation."""

import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from providers.excel import ExcelProvider
from providers.exceptions import InvalidFileStructureError


def _save_workbook(wb: Workbook) -> str:
    """Save workbook to a temporary file and return its path."""
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


@pytest.fixture()
def cleanup_files():
    """Collect temp file paths and remove them after the test."""
    paths: list[str] = []
    yield paths
    for p in paths:
        Path(p).unlink(missing_ok=True)


class TestPointsSheetValidation:
    """Tests for validation of the 'Punkty' sheet."""

    def test_missing_points_sheet(self, cleanup_files):
        """File without 'Punkty' sheet raises InvalidFileStructureError."""
        wb = Workbook()
        wb.active.title = "Inne dane"
        path = _save_workbook(wb)
        cleanup_files.append(path)

        provider = ExcelProvider(path)
        with pytest.raises(InvalidFileStructureError, match="Brak arkusza 'Punkty'"):
            provider.list_points()

    def test_missing_required_columns(self, cleanup_files):
        """'Punkty' sheet without required columns raises InvalidFileStructureError."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Punkty"
        ws.append(["Kolumna A", "Kolumna B"])
        path = _save_workbook(wb)
        cleanup_files.append(path)

        provider = ExcelProvider(path)
        with pytest.raises(InvalidFileStructureError, match="Brak wymaganych kolumn"):
            provider.list_points()

    def test_empty_points_sheet_returns_empty_list(self, cleanup_files):
        """'Punkty' sheet with headers but no data rows returns empty list."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Punkty"
        # Write all required column headers but no data rows
        headers = [
            "Nazwa punktu", "Kod punktu", "Współrzędne punktu",
            "Nazwa rzeki", "Kod JCWP", "Zarząd zlewni", "RZGW",
            "Opis lokalizacji", "Otoczenie", "Osoba badająca", "Kontakt",
        ]
        ws.append(headers)
        path = _save_workbook(wb)
        cleanup_files.append(path)

        provider = ExcelProvider(path)
        points = provider.list_points()
        assert points == []


class TestMeasurementSheetValidation:
    """Tests for validation of measurement sheets."""

    def _create_valid_points_file(self, point_name: str = "Punkt_1") -> Workbook:
        """Create a workbook with a valid 'Punkty' sheet referencing one point."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Punkty"
        headers = [
            "Nazwa punktu", "Kod punktu", "Współrzędne punktu",
            "Nazwa rzeki", "Kod JCWP", "Zarząd zlewni", "RZGW",
            "Opis lokalizacji", "Otoczenie", "Osoba badająca", "Kontakt",
        ]
        ws.append(headers)
        ws.append([point_name, "P1", "52.0 21.0", "Wisla", "J1", "Z1", "R1",
                    "Lokalizacja", "Las", "Jan", "jan@example.com"])
        return wb

    def test_missing_measurement_sheet(self, cleanup_files):
        """Missing measurement sheet for a point raises InvalidFileStructureError."""
        wb = self._create_valid_points_file("Punkt_1")
        # Do NOT create the 'Punkt_1' sheet
        path = _save_workbook(wb)
        cleanup_files.append(path)

        provider = ExcelProvider(path)
        points = provider.list_points()
        assert len(points) == 1

        with pytest.raises(InvalidFileStructureError, match="Brak arkusza 'Punkt_1'"):
            provider.list_measurements(points[0].id)

    def test_measurement_sheet_too_few_columns(self, cleanup_files):
        """Measurement sheet with fewer than 26 columns raises InvalidFileStructureError."""
        point_name = "Punkt_1"
        wb = self._create_valid_points_file(point_name)

        # Create measurement sheet with only 10 columns
        ws = wb.create_sheet(title=point_name)
        # Add header rows to skip (MEASUREMENTS_START_ROW = 8)
        for _ in range(8):
            ws.append(["header"] * 10)
        # Add a data row with only 10 columns
        ws.append(["data"] * 10)

        path = _save_workbook(wb)
        cleanup_files.append(path)

        provider = ExcelProvider(path)
        points = provider.list_points()

        with pytest.raises(
            InvalidFileStructureError, match=r"ma 10 kolumn, wymagane minimum 26"
        ):
            provider.list_measurements(points[0].id)
